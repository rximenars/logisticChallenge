import openpyxl
import numpy as np

archivoBogota = openpyxl.load_workbook ('Bogota.xlsx', data_only=True)
archivoMedellin = openpyxl.load_workbook ('Medellin.xlsx', data_only=True)
archivoBarranquilla = openpyxl.load_workbook ('Barranquilla.xlsx', data_only=True)
archivoBucaramanga = openpyxl.load_workbook ('Bucaramanga.xlsx', data_only=True)
archivoBogota = openpyxl.load_workbook ('Bogota.xlsx', data_only=True)
archivoBogota = openpyxl.load_workbook ('Bogota.xlsx', data_only=True)


bogota = archivoBogota.active 

celdas = bogota['A3' :'A224684']


   
def extraerBase (celda):

    mayuscula = celda.upper()
    

    primer_base = [
               ["AVENIDA CARRERA","AVENIDA CARRERA"],
               ["AVENIDA CARRERA","AK"],
               ["AVENIDA CARRERA","A.K"],
               ["AVENIDA CARRERA","AV CRR"],
               ["AVENIDA CARRERA","AV. CRR"],
               ["AVENIDA CARRERA","AV CRA"],
               ["AVENIDA CARRERA","AV. CRA"],
               ["AVENIDA CALLE","AVENIDA CALLE"],
               ["AVENIDA CALLE","AC"],
               ["AVENIDA CALLE","A.C"],
               ["AVENIDA CALLE","AV. CLL"],
               ["AVENIDA CALLE","AV. CALLE"],
               ["AVENIDA CALLE","AV CLL"],
               ["AVENIDA CALLE","AV CALLE"],
               ["CIRCUNVALAR","CIRCUNVALAR"],
               ["CIRCUNVALAR","AV. CIRCUNVALAR"],
               ["CIRCUNVALAR","AV CIRCUNVALAR"],
               ["CIRCUNVALAR","A CIRCUNVALAR"],
               ["CALLE","CALLE"],
               ["CALLE", "CLL"],
               ["CALLE","CL"],
               ["CARRERA","CARRERA"],
               ["CARRERA","KRA"],
               ["CARRERA","CRA"],
               ["CARRERA","CRR"],
               ["CARRERA","KRR"],
               ["CARRERA","CR"],
               ["CARRERA","KR"],
               ["AVENIDA","AVENIDA"],
               ["AVENIDA","AVNDA"],
               ["AVENIDA","AVDA"],
               ["AVENIDA","AVE"],
               ["AVENIDA","AV"],
               ["DIAGONAL","DIAGONAL"],
               ["DIAGONAL","DIAG"],
               ["DIAGONAL","DIG"],
               ["DIAGONAL","DG"],
               ["MANZANA","MANZANA"],
               ["MANZANA","MN"],
               ["MANZANA","MZN"],
               ["TRANSVERSAL","TRANSVERSAL"],
               ["TRANSVERSAL","TRANSV"],
               ["TRANSVERSAL","TRANS"],
               ["TRANSVERSAL","TRAS"],
               ["TRANSVERSAL","TRAN"],
               ["TRANSVERSAL","TR"],
               ["TRANSVERSAL","TN"],
               ["TRANSVERSAL","TV"],
               ["VIA","VIA"]
               ]

    segunda_base = [
                ["#","NO"],
                ["#","NUMERO"]
                ]


    tercera_base = [
                ["-","-"],
                [" "," "]
                ] 


    prefijos = np.unique(np.array(primer_base)[:,0])
    arrayCheck = True

    inicio_cadenaTexto = 0
    fin_cadenaTexto = 0

    for x in primer_base:
        if(mayuscula.find(x[1])>=fin_cadenaTexto):
            mayuscula = mayuscula.replace(x[1],x[0])
            fin_cadenaTexto = mayuscula.find (x[0])+len(x[0])
            

    
    for x in segunda_base:
        if((mayuscula.find(x[1],fin_cadenaTexto)>=fin_cadenaTexto) & arrayCheck):
            mayuscula = mayuscula.replace(x[1],x[0])
            fin_cadenaTexto+= mayuscula.find (x[0],fin_cadenaTexto)+len(x[0])
            arrayCheck = False
     
    for x in tercera_base:
        if(mayuscula.find(x[1])>=fin_cadenaTexto):
            fin_cadenaTexto+= mayuscula.find (x[0],fin_cadenaTexto)+len(x[0])


    for x in prefijos:
        if((mayuscula.find(x)>=inicio_cadenaTexto)):
            inicio_cadenaTexto = mayuscula.find(x) 

    recortar = mayuscula [inicio_cadenaTexto:fin_cadenaTexto]
    complemento = mayuscula.replace(recortar,"")
    return [recortar,complemento]


for fila in celdas:
    print ([extraerBase(celda.value) for celda in fila])    


    
    
    